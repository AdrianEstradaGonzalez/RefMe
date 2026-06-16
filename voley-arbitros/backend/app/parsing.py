"""Lectura de los ficheros Excel de origen (árbitros, calendario, categorías).

Estas funciones son "puras": sólo dependen de openpyxl y devuelven listas de
diccionarios. No tocan la base de datos, lo que permite probarlas de forma
aislada y reutilizarlas tanto en la carga inicial (`seed`) como en la
importación desde la interfaz web.
"""
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Jerarquía de niveles arbitrales, de menor a mayor competencia.
# El índice (empezando en 1) se usa como campo `orden`.
NIVELES_ORDENADOS = [
    "Candidato Territorial I Pista",
    "Nivel I Pista",
    "Nivel I + Hab. Nivel II Pista",
    "Nivel II Pista",
    "Nivel II + Hab. Nacional C Pista",
    "Nacional C Pista",
    "Nacional B Pista",
    "Nacional A Pista",
    "Internacional Pista",
]

FRANJAS = ["09:00-12:00", "12:00-15:00", "15:00-18:00", "18:00-22:00"]


def _limpiar(valor):
    """Normaliza una celda de texto: quita espacios raros (incl. \xa0)."""
    if valor is None:
        return None
    texto = str(valor).replace("\xa0", " ").strip()
    return texto or None


def _parse_fecha(valor):
    """Convierte 'DD/MM/YYYY' (o un datetime) a date."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _parse_hora(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")
    texto = str(valor).strip()
    return texto or None


def _to_int(valor):
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def leer_arbitros(ruta: Path):
    """Devuelve [{'nombre', 'codigo'}] desde la hoja 'Arbitros'."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Arbitros"]
    salida = []
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not fila or not fila[0]:
            continue
        salida.append({"nombre": _limpiar(fila[0]), "codigo": _to_int(fila[1])})
    wb.close()
    return salida


def leer_categorias(ruta: Path):
    """Lee la hoja 'Hoja1' del fichero de categorías/niveles.

    Columnas: Categoria | Primero | Segundo | Anotador | Nº Mínimo | Nivel | Prioridad
    El valor "NO" en un rol significa que ese rol no se exige.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Hoja1"]
    salida = []
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not fila or not fila[0]:
            continue
        salida.append(
            {
                "nombre": _limpiar(fila[0]),
                "nivel_primero": _nivel_rol(fila[1]),
                "nivel_segundo": _nivel_rol(fila[2]),
                "nivel_anotador": _nivel_rol(fila[3]),
                "min_arbitros": _to_int(fila[4]) or 1,
                "nivel_general": _nivel_rol(fila[5]),
                "prioridad": _to_int(fila[6]),
            }
        )
    wb.close()
    return salida


def _nivel_rol(valor):
    """Devuelve el nombre de nivel normalizado o None si es 'NO'/vacío."""
    texto = _limpiar(valor)
    if not texto or texto.upper() == "NO":
        return None
    return texto


def leer_equipos(ruta: Path):
    """Devuelve [{'categoria', 'equipo', 'club'}] desde la hoja 'Equipos'."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Equipos"]
    salida = []
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not fila or not fila[0]:
            continue
        salida.append(
            {
                "categoria": _limpiar(fila[0]),
                "equipo": _limpiar(fila[1]),
                "club": _limpiar(fila[2]),
            }
        )
    wb.close()
    return salida


def leer_calendario(ruta: Path):
    """Lee la hoja 'Calendario' y devuelve partidos + árbitros ya designados.

    Cada elemento incluye los nombres de Árbitro 1 / Árbitro 2 / Anotador tal
    como aparecen, para poder reconstruir asignaciones de ejemplo.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Calendario"]
    salida = []
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not fila or not fila[2]:
            continue
        salida.append(
            {
                "fecha": _parse_fecha(fila[0]),
                "hora": _parse_hora(fila[1]),
                "categoria": _limpiar(fila[2]),
                "jornada": _to_int(fila[3]),
                "numero_partido": _to_int(fila[4]),
                "local": _limpiar(fila[5]),
                "visitante": _limpiar(fila[6]),
                "arbitro1": _nombre_arbitro(fila[7]),
                "arbitro2": _nombre_arbitro(fila[8]),
                "anotador": _nombre_arbitro(fila[9]),
                "campo": _limpiar(fila[10]),
                "provincia": _limpiar(fila[11]),
                "codigo_partido": _to_int(fila[12]),
            }
        )
    wb.close()
    return salida


def _nombre_arbitro(valor):
    """Filtra valores que no son árbitros reales (p. ej. 'Por Determinar')."""
    texto = _limpiar(valor)
    if not texto or texto.lower().startswith("por determinar"):
        return None
    return texto


# --------------------------------------------------------------- Licencias
# El fichero "Relacion_licencias" agrupa a los árbitros por nivel mediante
# filas-cabecera (sólo la columna A rellena) y filas de columnas ("LICENCIA").
# En las filas de datos las columnas de domicilio aparecen DESPLAZADAS respecto
# a su cabecera: la dirección está en la col. J, el código postal en la K y el
# municipio en la L.
def leer_licencias(ruta: Path):
    """Devuelve [{'nombre', 'direccion', 'codigo_postal', 'localidad'}].

    El nombre se reconstruye como "NOMBRE APELLIDOS" para poder cruzarlo con el
    cuerpo arbitral ya cargado. Se descartan cabeceras y filas duplicadas.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Listado Licencias"]
    salida = []
    vistos = set()
    for fila in ws.iter_rows(values_only=True):
        col_a = _limpiar(fila[0])
        nombre = _limpiar(fila[3])
        # Saltar cabeceras de sección y la fila de títulos de columna.
        if not col_a or not nombre or col_a == "LICENCIA":
            continue
        completo = f"{nombre} {_limpiar(fila[4]) or ''}".strip()
        clave = completo.upper()
        if clave in vistos:
            continue
        vistos.add(clave)
        salida.append(
            {
                "nombre": completo,
                "direccion": _limpiar(fila[9]),
                "codigo_postal": _codigo_postal(fila[10]),
                "localidad": _limpiar(fila[11]),
            }
        )
    wb.close()
    return salida


def _codigo_postal(valor):
    cp = _to_int(valor)
    return f"{cp:05d}" if cp is not None else None


# ------------------------------------------------------------ Disponibilidad
# Año de la temporada al que corresponden las columnas de fecha (formato DD-MM).
# Verificado contra el calendario: 01-01 cae en miércoles → 2025.
ANIO_DISPONIBILIDAD = 2025


def leer_disponibilidad(ruta: Path):
    """Lee el grid de disponibilidad por árbitro, fecha y franja.

    Estructura del fichero: cada árbitro ocupa 5 filas (4 franjas horarias en el
    orden de `FRANJAS` + 1 fila en blanco de separación). La fila de identidad
    lleva nombre (col. B), ciudad (col. D) y coche (col. E) y, además, los datos
    de la primera franja. Las columnas F en adelante son días "DD-MM".

    Devuelve [{'nombre', 'ciudad', 'coche': bool,
               'celdas': [(date, franja, 'SI'|'NO'), ...]}].
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["Disponibilidad"]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    cabecera = filas[0]
    fechas = [_fecha_columna(c) for c in cabecera[5:]]

    salida = []
    i = 2  # las dos primeras filas son cabecera + letras de día de la semana
    while i < len(filas):
        ident = filas[i]
        nombre = _limpiar(ident[1])
        if not nombre:
            i += 1
            continue

        bloque = filas[i : i + len(FRANJAS)]
        celdas = []
        for franja, fila in zip(FRANJAS, bloque):
            for fecha, celda in zip(fechas, fila[5:]):
                estado = _limpiar(celda)
                if fecha is None or estado not in ("SI", "NO"):
                    continue
                celdas.append((fecha, franja, estado))

        salida.append(
            {
                "nombre": nombre,
                "nivel": _limpiar(ident[0]),  # col A: clasificación arbitral
                "ciudad": _limpiar(ident[3]),
                "coche": _tiene_coche(ident[4]),
                "celdas": celdas,
            }
        )
        i += len(FRANJAS) + 1  # saltar el bloque + la fila en blanco

    return salida


def _fecha_columna(valor):
    """Convierte el encabezado 'DD-MM' a un date del año de la temporada."""
    texto = _limpiar(valor)
    if not texto:
        return None
    try:
        dia, mes = (int(p) for p in texto.split("-"))
        return date(ANIO_DISPONIBILIDAD, mes, dia)
    except (ValueError, TypeError):
        return None


def _tiene_coche(valor):
    """'Sí' → True; 'No' / 'No Indicado' / 'Preguntar' / vacío → False."""
    texto = _limpiar(valor)
    return bool(texto) and texto.upper().startswith("S")
